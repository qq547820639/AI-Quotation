"""询价单路由：list / get / create / update / delete + 6 动作端点 + 按 inquiry 查报价

动作端点均：更新 status + 追加 InquiryLog + 更新 updatedAt + 审批端点追加 ApprovalNode。
对齐 src/mocks/handlers.ts 业务逻辑。
"""
from __future__ import annotations
from typing import Optional

import json
import secrets
from datetime import datetime

from fastapi import APIRouter, Body, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload, selectinload

from ..database import get_db
from .. import config
from ..invitations import regenerate_invitation, resend_invitation, InvitationError
from ..delivery import (
    ensure_invitations, delivery_summary,
    generate_deadline_reminders,
)
from ..models import (
    User, Inquiry, InquiryItem, InquiryLog, ApprovalNode, Quotation,
    Supplier, AppSettings, SupplierInvitation, QuotationSnapshot,
)
from ..schemas import (
    InquirySchema, InquiryCreate, InquiryUpdate, ApprovalAction,
    QuotationSchema, SuccessResult, VersionBody,
    DeliveryRecordSchema, DeliverySummarySchema,
    PaginatedInquiriesSchema, ExportRequest, QuotationSnapshotSchema,
)
from ..auth import get_current_user, require_permission, resolve_permissions
from ..serializers import inquiry_to_schema, quotation_to_schema, gen_id, now_str
from ..policy import require_inquiry_access, require_inquiry_edit, filter_visible_inquiries, set_create_ownership
from ..state_machine import (
    assert_inquiry_transition, S_PENDING_SEND, S_INQUIRING, S_CANCELLED, S_COMPLETED,
    S_PENDING_APPROVAL, S_PENDING_CONFIRM,
)
from ..templates import preview_template
from ..events import publish

router = APIRouter(prefix="/inquiries", tags=["inquiries"])

# 状态枚举值（对齐前端 InquiryStatus）
S_PENDING_APPROVAL = "PENDING_APPROVAL"
S_PENDING_CONFIRM = "PENDING_CONFIRM"
S_RETURNED = "RETURNED"
S_INQUIRING = "INQUIRING"
S_COMPLETED = "COMPLETED"
S_CANCELLED = "CANCELLED"

# 审批节点状态
APV_PENDING = "PENDING"
APV_APPROVED = "APPROVED"
APV_REJECTED = "REJECTED"

# 日志类型
LOG_SEND_INQUIRY = "SEND_INQUIRY"
LOG_CANCEL = "CANCEL"
LOG_CONFIRM_RESULT = "CONFIRM_RESULT"
LOG_SUBMIT_APPROVAL = "SUBMIT_APPROVAL"
LOG_APPROVE = "APPROVE"
LOG_REJECT = "REJECT"


def _append_log(db: Session, inquiry: Inquiry, user: User, log_type: str, content: str, result: str | None = None):
    log = InquiryLog(
        id=gen_id(f"log-{inquiry.id}"),
        inquiry_id=inquiry.id,
        time=now_str(),
        operator=user.name,
        operator_role=user.role,
        type=log_type,
        content=content,
        result=result,
    )
    db.add(log)
    inquiry.logs.append(log)


def _verify_version(inquiry: Inquiry, body_version: int | None) -> None:
    """乐观锁校验：客户端携带的 version 与当前不一致时返回 409（Task 6）"""
    if body_version is not None and inquiry.version != body_version:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="数据已被他人修改，请刷新后重试",
        )


def _generate_inquiry_code(db: Session) -> str:
    """生成唯一询价编号：INQ + YYYYMMDD + 3 位随机序号（Task 7）

    数据库 code 列有唯一约束，最大重试避免碰撞；碰撞由调用方在事务内整体重试。
    """
    base = datetime.now().strftime("%Y%m%d")
    for _ in range(50):
        code = f"INQ{base}{secrets.randbelow(999) + 1:03d}"
        if db.query(Inquiry).filter(Inquiry.code == code).first() is None:
            return code
    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail="无法生成唯一询价编号")


def _merge_map(db: Session, inquiry: Inquiry, key: str, incoming: dict | None) -> None:
    """增量合并 JSON 字段（selected_supplier_map / purchaser_comments），避免覆盖其他键（Task 5）"""
    if not incoming:
        return
    current = dict(getattr(inquiry, key) or {})
    current.update(incoming)
    setattr(inquiry, key, current)


def _build_inquiry_items(inquiry_id: str, items_data: list) -> list[InquiryItem]:
    """从前端 items 构造 ORM InquiryItem 列表（items 含 material 内联对象 + 扁平字段）"""
    result = []
    for it in items_data or []:
        material = it.get("material") or {}
        result.append(InquiryItem(
            id=it.get("id") or gen_id(f"item-{inquiry_id}"),
            inquiry_id=inquiry_id,
            material_id=material.get("id") or it.get("materialId"),
            name=it.get("name", material.get("name", "")),
            code=it.get("code", material.get("code", "")),
            category=it.get("category", material.get("category", "")),
            brand=it.get("brand", material.get("brand", "")),
            spec=it.get("spec", material.get("spec", "")),
            tech_params=it.get("techParams", material.get("techParams", "")),
            unit=it.get("unit", material.get("unit", "")),
            quantity=it.get("quantity", 0),
            target_price=it.get("targetPrice"),
            expected_delivery_date=it.get("expectedDeliveryDate"),
            remark=it.get("remark"),
        ))
    return result


def _build_logs(inquiry_id: str, logs_data: list, default_user: User | None = None) -> list[InquiryLog]:
    """从前端 logs 构造 ORM InquiryLog 列表"""
    result = []
    for lg in logs_data or []:
        result.append(InquiryLog(
            id=lg.get("id") or gen_id(f"log-{inquiry_id}"),
            inquiry_id=inquiry_id,
            time=lg.get("time", now_str()),
            operator=lg.get("operator", default_user.name if default_user else "系统"),
            operator_role=lg.get("operatorRole", default_user.role if default_user else "系统"),
            type=lg.get("type", "CREATE"),
            content=lg.get("content", ""),
            result=lg.get("result"),
        ))
    return result


def _build_approval_nodes(inquiry_id: str, nodes_data: list) -> list[ApprovalNode]:
    result = []
    for nd in nodes_data or []:
        result.append(ApprovalNode(
            id=nd.get("id") or gen_id(f"apv-{inquiry_id}"),
            inquiry_id=inquiry_id,
            node_order=nd.get("nodeOrder", 1),
            approver_id=nd.get("approverId", ""),
            approver_name=nd.get("approverName", ""),
            approver_role=nd.get("approverRole", ""),
            status=nd.get("status", APV_PENDING),
            comment=nd.get("comment"),
            time=nd.get("time"),
        ))
    return result


# ============ 查询 ============

# 排序白名单：仅允许按固定字段排序，避免 SQL 注入/任意列排序
_SORT_FIELDS = {
    "updatedAt": Inquiry.updated_at,
    "createdAt": Inquiry.created_at,
    "deadline": Inquiry.deadline,
    "code": Inquiry.code,
    "subject": Inquiry.subject,
}


@router.get("")
def list_inquiries(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    page: Optional[int] = Query(default=None, ge=1),
    pageSize: Optional[int] = Query(default=None, ge=1, le=200),
    keyword: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    dateFrom: Optional[str] = Query(default=None),
    dateTo: Optional[str] = Query(default=None),
    sort: Optional[str] = Query(default=None),
):
    """询价列表（P2-12 Task 17 服务端分页/筛选/搜索/排序）

    - 向后兼容：不传分页参数时返回全量数组（保持既有调用不变）。
    - 传入 page/pageSize 时返回分页结构 {items, total, page, pageSize}。
    - keyword 匹配 code/subject/owner_name；status 为逗号分隔的状态列表；
      dateFrom/dateTo 过滤 created_at（YYYY-MM-DD）；sort 如 "updatedAt:desc"。
    """
    query = db.query(Inquiry)
    query = filter_visible_inquiries(query, user)
    # P2 Task 22：selectinload 预加载集合关系，消除列表页 N+1
    # （items / quotations(含 items) / logs / approval_nodes / invited_suppliers）
    query = query.options(
        selectinload(Inquiry.items),
        selectinload(Inquiry.quotations).selectinload(Quotation.items),
        selectinload(Inquiry.logs),
        selectinload(Inquiry.approval_nodes),
        selectinload(Inquiry.invited_suppliers),
    )

    # 关键词搜索（code / subject / owner_name）
    if keyword:
        kw = f"%{keyword.strip()}%"
        query = query.filter(or_(
            Inquiry.code.like(kw),
            Inquiry.subject.like(kw),
            Inquiry.owner_name.like(kw),
        ))
    # 状态筛选
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        if statuses:
            query = query.filter(Inquiry.status.in_(statuses))
    # 创建时间范围
    if dateFrom:
        query = query.filter(Inquiry.created_at >= dateFrom)
    if dateTo:
        query = query.filter(Inquiry.created_at <= dateTo)

    # 排序（白名单 + 方向）
    order_col = _SORT_FIELDS.get(sort.split(":")[0]) if sort else Inquiry.updated_at
    direction = sort.split(":", 1)[1] if sort and ":" in sort else "desc"
    col = order_col.asc() if direction == "asc" else order_col.desc()
    # Task 7：id 作为稳定次排序键，避免分页边界重复/漏行
    query = query.order_by(col, Inquiry.id)

    # 分页：未传分页参数则返回全量数组（向后兼容）
    if page is None and pageSize is None:
        rows = query.all()
        return [inquiry_to_schema(i, db) for i in rows]

    _page = page if page is not None else 1
    _size = pageSize if pageSize is not None else 20
    total = query.count()
    rows = query.offset((_page - 1) * _size).limit(_size).all()
    return PaginatedInquiriesSchema(
        items=[inquiry_to_schema(i, db) for i in rows],
        total=total,
        page=_page,
        pageSize=_size,
    )


@router.get("/{inquiry_id}/quotations", response_model=list[QuotationSchema])
def list_quotations_by_inquiry(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    rows = (
        db.query(Quotation)
        .options(joinedload(Quotation.items))
        .filter(Quotation.inquiry_id == inquiry_id)
        .order_by(Quotation.created_at, Quotation.id)
        .all()
    )
    return [quotation_to_schema(q, db) for q in rows]


@router.get("/{inquiry_id}", response_model=InquirySchema)
def get_inquiry(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    return inquiry_to_schema(inq, db)


# ============ 增删改 ============

@router.post("", response_model=InquirySchema)
def create_inquiry(
    body: InquiryCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_CREATE")),
):
    data = body.model_dump()
    inq_id = data.get("id") or gen_id("inq")
    ts = data.get("createdAt") or now_str()
    # 强制以当前登录用户作为组织/所有者/创建者，忽略客户端提交的值（P0-资源级授权）
    set_create_ownership(data, user)

    def build_inquiry(code: str) -> Inquiry:
        return Inquiry(
            id=inq_id,
            code=code,
            subject=data.get("subject", ""),
            organization=data["organization"],
            owner_name=data["owner_name"],
            owner_id=data["owner_id"],
            currency=data.get("currency", "CNY"),
            deadline=data.get("deadline", ""),
            expected_delivery_date=data.get("expectedDeliveryDate"),
            delivery_address=data.get("deliveryAddress", ""),
            contact=data.get("contact", ""),
            payment_terms=data.get("paymentTerms", ""),
            invoice_requirement=data.get("invoiceRequirement"),
            description=data.get("description"),
            status=data.get("status", "DRAFT"),
            created_by_id=data["created_by_id"],
            created_by_name=data["created_by_name"],
            created_at=ts,
            updated_at=data.get("updatedAt") or ts,
            selected_supplier_map=data.get("selectedSupplierMap", {}),
            purchaser_comments=data.get("purchaserComments", {}),
            version=1,
        )

    # 服务端生成唯一编号（Task 7）：并发碰撞时事务内整体重试
    for _ in range(5):
        inq = build_inquiry(_generate_inquiry_code(db))
        inq.items = _build_inquiry_items(inq_id, data.get("items", []))
        inq.logs = _build_logs(inq_id, data.get("logs", []), default_user=user)
        inq.approval_nodes = _build_approval_nodes(inq_id, data.get("approvalNodes", []))
        for sup_id in data.get("invitedSupplierIds", []) or []:
            sup = db.query(Supplier).filter(Supplier.id == sup_id).first()
            if sup is not None:
                inq.invited_suppliers.append(sup)
        db.add(inq)
        try:
            db.commit()
            db.refresh(inq)
            return inquiry_to_schema(inq, db)
        except IntegrityError:
            db.rollback()
        except Exception:
            db.rollback()
            raise
    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail="创建询价单失败：编号生成冲突重试耗尽")


@router.put("/{inquiry_id}", response_model=InquirySchema)
def update_inquiry(
    inquiry_id: str,
    body: InquiryUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_EDIT")),
):
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_edit(user, inq)
    data = body.model_dump(exclude_unset=True)
    # Task 6：乐观锁校验
    _verify_version(inq, data.get("version"))
    # 标量字段：不允许修改 status / organization / createdById / createdByName / code
    scalar_map = {
        "subject": "subject", "ownerName": "owner_name", "ownerId": "owner_id",
        "currency": "currency", "deadline": "deadline",
        "expectedDeliveryDate": "expected_delivery_date",
        "deliveryAddress": "delivery_address", "contact": "contact",
        "paymentTerms": "payment_terms", "invoiceRequirement": "invoice_requirement",
        "description": "description",
    }
    for camel, snake in scalar_map.items():
        if camel in data:
            setattr(inq, snake, data[camel])
    # Task 5：JSON 字段增量合并，避免覆盖其他物料/供应商键
    if "selectedSupplierMap" in data:
        _merge_map(db, inq, "selected_supplier_map", data["selectedSupplierMap"])
    if "purchaserComments" in data:
        _merge_map(db, inq, "purchaser_comments", data["purchaserComments"])
    # items（整体替换）
    if "items" in data:
        for old in inq.items:
            db.delete(old)
        inq.items = _build_inquiry_items(inq.id, data["items"])
    # invited suppliers（整体替换）
    if "invitedSupplierIds" in data:
        inq.invited_suppliers = []
        for sup_id in data["invitedSupplierIds"] or []:
            sup = db.query(Supplier).filter(Supplier.id == sup_id).first()
            if sup is not None:
                inq.invited_suppliers.append(sup)
    inq.updated_at = now_str()
    inq.version += 1  # Task 6：成功更新后递增版本号
    db.commit()
    db.refresh(inq)
    return inquiry_to_schema(inq, db)


@router.delete("/{inquiry_id}", response_model=SuccessResult)
def delete_inquiry(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_CANCEL")),
):
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_edit(user, inq)
    db.delete(inq)
    db.commit()
    return SuccessResult(success=True)


# ============ 动作端点 ============

@router.post("/{inquiry_id}/send", response_model=InquirySchema)
def send_inquiry(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_SEND")),
    body: VersionBody = Body(default=None),
):
    """发送询价（P1-8 Task 12）：校验状态机 → 为受邀供应商创建投递记录 →
    通过持久化队列（outbox + Celery）触发异步投递 → 状态流转 INQUIRING。

    投递结果由 GET /deliveries 反映真实交付状态，不在此处谎报"已全部发送成功"。
    队列入队幂等（同 inquiry 重复发送被跳过），eager 模式下同步执行保证测试不失真。
    """
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    _verify_version(inq, body.version if body else None)
    assert_inquiry_transition(inq.status, S_INQUIRING)
    count = len(inq.invited_suppliers)
    # 为每个受邀供应商创建/更新投递记录（幂等）
    ensure_invitations(db, inq, user)
    inq.status = S_INQUIRING
    _append_log(db, inq, user, LOG_SEND_INQUIRY, f"向 {count} 家供应商发送询价", "询价中")
    inq.updated_at = now_str()
    inq.version += 1
    db.commit()
    # 通过持久化队列投递邮件（独立事务 outbox；无渠道则投递记录保持 pending）
    from ..queue_client import enqueue
    enqueue("email.send", inquiry_id, {"inquiry_id": inquiry_id})
    db.refresh(inq)
    return inquiry_to_schema(inq, db)


@router.post("/{inquiry_id}/cancel", response_model=InquirySchema)
def cancel_inquiry(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_CANCEL")),
    body: VersionBody = Body(default=None),
):
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    _verify_version(inq, body.version if body else None)
    assert_inquiry_transition(inq.status, S_CANCELLED)
    inq.status = S_CANCELLED
    _append_log(db, inq, user, LOG_CANCEL, "取消询价单", "已取消")
    inq.updated_at = now_str()
    inq.version += 1
    db.commit()
    db.refresh(inq)
    return inquiry_to_schema(inq, db)


@router.post("/{inquiry_id}/confirm", response_model=InquirySchema)
def confirm_inquiry(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_CONFIRM")),
    body: VersionBody = Body(default=None),
):
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    _verify_version(inq, body.version if body else None)
    assert_inquiry_transition(inq.status, S_COMPLETED)
    inq.status = S_COMPLETED
    _append_log(db, inq, user, LOG_CONFIRM_RESULT, "确认定标结果", "已完成")
    inq.updated_at = now_str()
    inq.version += 1
    # P2-12 Task 17：定标时冻结报价为不可变快照，避免供应商后续修改影响审批记录
    snapshot = _create_quotation_snapshot(db, inq, user)
    db.add(snapshot)
    db.commit()
    db.refresh(inq)
    publish("inquiry_confirmed", {"inquiryId": inq.id, "code": inq.code})
    return inquiry_to_schema(inq, db)


@router.post("/{inquiry_id}/submit-approval", response_model=InquirySchema)
def submit_approval(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_SEND")),
    body: VersionBody = Body(default=None),
):
    """提交审批：status→PENDING_APPROVAL + 新增 PENDING 审批节点 + 追加 SUBMIT_APPROVAL 日志

    审批人取自 AppSettings.approval_approver_id（默认 u-2）
    """
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    _verify_version(inq, body.version if body else None)
    assert_inquiry_transition(inq.status, S_PENDING_APPROVAL)
    # 读取审批配置
    settings = db.query(AppSettings).filter(AppSettings.id == 1).first()
    approver_id = settings.approval_approver_id if settings else "u-2"
    approver = db.query(User).filter(User.id == approver_id).first()
    approver_name = approver.name if approver else "采购主管"
    approver_role = approver.role if approver else "采购主管"
    node = ApprovalNode(
        id=gen_id(f"apv-{inq.id}"),
        inquiry_id=inq.id,
        node_order=len(inq.approval_nodes) + 1,
        approver_id=approver_id,
        approver_name=approver_name,
        approver_role=approver_role,
        status=APV_PENDING,
    )
    db.add(node)
    inq.approval_nodes.append(node)
    inq.status = S_PENDING_APPROVAL
    _append_log(db, inq, user, LOG_SUBMIT_APPROVAL, f"提交审批，审批人：{approver_name}")
    inq.updated_at = now_str()
    inq.version += 1
    db.commit()
    db.refresh(inq)
    return inquiry_to_schema(inq, db)


@router.post("/{inquiry_id}/approve", response_model=InquirySchema)
def approve_inquiry(
    inquiry_id: str,
    body: ApprovalAction,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_APPROVE")),
):
    """审批通过：status→PENDING_CONFIRM + PENDING 节点转 APPROVED + 追加 APPROVE 日志"""
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    _verify_version(inq, body.version)
    assert_inquiry_transition(inq.status, S_PENDING_CONFIRM)
    ts = now_str()
    for node in inq.approval_nodes:
        if node.status == APV_PENDING:
            node.status = APV_APPROVED
            node.comment = body.comment
            node.time = ts
    inq.status = S_PENDING_CONFIRM
    comment_suffix = f"：{body.comment}" if body.comment else ""
    _append_log(db, inq, user, LOG_APPROVE, f"审批通过{comment_suffix}", "已通过")
    inq.updated_at = ts
    inq.version += 1
    db.commit()
    db.refresh(inq)
    return inquiry_to_schema(inq, db)


@router.post("/{inquiry_id}/reject", response_model=InquirySchema)
def reject_inquiry(
    inquiry_id: str,
    body: ApprovalAction,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_APPROVE")),
):
    """审批驳回：status→RETURNED（可重新编辑） + PENDING 节点转 REJECTED + 追加 REJECT 日志"""
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    _verify_version(inq, body.version)
    assert_inquiry_transition(inq.status, S_RETURNED)
    ts = now_str()
    for node in inq.approval_nodes:
        if node.status == APV_PENDING:
            node.status = APV_REJECTED
            node.comment = body.comment
            node.time = ts
    inq.status = S_RETURNED
    comment_suffix = f"：{body.comment}" if body.comment else ""
    _append_log(db, inq, user, LOG_REJECT, f"审批驳回{comment_suffix}", "已驳回")
    inq.updated_at = ts
    inq.version += 1
    db.commit()
    db.refresh(inq)
    return inquiry_to_schema(inq, db)


# ============ P1-8 Task 12：交付状态 / 重发 / 截止提醒 / 模板预览 ============

def _delivery_record_to_dict(inv: SupplierInvitation, db: Session) -> dict:
    supplier = db.query(Supplier).filter(Supplier.id == inv.supplier_id).first()
    return {
        "supplierId": inv.supplier_id,
        "supplierName": supplier.name if supplier else "",
        "deliveryStatus": inv.delivery_status,
        "invitationStatus": inv.status,
        "sentAt": inv.sent_at.isoformat() if inv.sent_at else None,
        "openedAt": inv.first_opened_at.isoformat() if inv.first_opened_at else None,
        "submittedAt": inv.submitted_at.isoformat() if inv.submitted_at else None,
        "deliveryError": inv.delivery_error,
    }


@router.get("/templates/preview")
def preview_inquiry_template(
    body: dict = Body(default=None),
    _: User = Depends(require_permission("INQUIRY_SEND")),
):
    """询价模板预览：返回渲染后的主题/正文与缺失变量（支持多语言）。"""
    variables = (body or {}).get("variables", {})
    lang = (body or {}).get("lang")
    return preview_template("inquiry", lang, variables)


@router.post("/reminders/deadline")
def trigger_deadline_reminders(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("INQUIRY_SEND")),
):
    """为临近截止的未提交供应商生成提醒通知（幂等），返回本次创建条数。"""
    created = generate_deadline_reminders(db)
    return {"created": created}


@router.get("/{inquiry_id}/deliveries")
def get_inquiry_deliveries(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """返回逐供应商交付状态 + 汇总（采购端查看真实投递结果）。"""
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    invitations = db.query(SupplierInvitation).filter(
        SupplierInvitation.inquiry_id == inquiry_id,
    ).order_by(SupplierInvitation.supplier_id).all()
    suppliers = [_delivery_record_to_dict(i, db) for i in invitations]
    return {
        "suppliers": suppliers,
        "summary": delivery_summary(invitations),
    }


@router.post("/{inquiry_id}/deliveries/{supplier_id}/resend")
def resend_inquiry_delivery(
    inquiry_id: str,
    supplier_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_SEND")),
):
    """重新发送给指定供应商。

    复用短期存储中仍有效的原始 token 重新投递；若明文 token 已丢失（Redis 重启/进程重启）
    则自动重签新 token 并投递，保证重发的一定是有效链接。已撤销/已提交/已过期/询价已终结的
    邀请重发返回结构化错误（409/410），不静默。
    """
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    inv = db.query(SupplierInvitation).filter(
        SupplierInvitation.inquiry_id == inquiry_id,
        SupplierInvitation.supplier_id == supplier_id,
    ).first()
    if inv is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="该供应商的邀请不存在")
    try:
        resend_invitation(db, inv, user.id)
    except InvitationError as e:
        raise HTTPException(
            status_code=e.status_code,
            detail={"error_type": e.error_type, "message": e.message},
        )
    # 通过持久化队列重投（显式唯一幂等键，保证每次重发都真正投递）
    from ..queue_client import enqueue
    enqueue("email.send", inquiry_id, {"inquiry_id": inquiry_id},
            idempotency_key=f"resend:{inv.id}:{now_str()}")
    return _delivery_record_to_dict(inv, db)


@router.post("/{inquiry_id}/invitations/{supplier_id}/regenerate")
def regenerate_invitation_link(
    inquiry_id: str,
    supplier_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("INQUIRY_SEND")),
):
    """重新生成某供应商的邀请链接（P0-2 撤销/重新生成能力，也为采购端提供可分享链接）。

    撤销旧邀请并创建新邀请，返回新的原始 token 与门户链接。原始 token 仅经短期存储
    返回一次，不落库；门户侧按 token 哈希校验，无法通过枚举 ID 或 token 访问他人资源。
    若供应商不在该询价受邀名单中则拒绝（资源级校验）。
    """
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    invited_ids = {s.id for s in inq.invited_suppliers}
    if supplier_id not in invited_ids:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该供应商未被邀请到本询价单")
    inv = db.query(SupplierInvitation).filter(
        SupplierInvitation.inquiry_id == inquiry_id,
        SupplierInvitation.supplier_id == supplier_id,
    ).first()
    if inv is None:
        # 尚未创建邀请（如询价未发送）：先创建，再返回其链接
        from ..invitations import create_invitation
        raw_token, inv = create_invitation(db, inquiry_id, supplier_id, user.id)
    else:
        raw_token, inv = regenerate_invitation(db, inv, user.id)
    return {
        "token": raw_token,
        "invitationId": inv.id,
        "supplierId": supplier_id,
        "inquiryId": inquiry_id,
        "portalUrl": config.build_invitation_url(raw_token),
    }


# ============ P2-12 Task 17：报价快照 / 服务端导出 ============

def _create_quotation_snapshot(db: Session, inq: Inquiry, user: User | None) -> QuotationSnapshot:
    """将询价摘要 + 全部已提交报价冻结为不可变 JSON 快照。"""
    quotations = db.query(Quotation).filter(Quotation.inquiry_id == inq.id).all()
    frozen = {
        "inquiry": {
            "id": inq.id,
            "code": inq.code,
            "subject": inq.subject,
            "currency": inq.currency,
            "status": inq.status,
            "selectedSupplierMap": inq.selected_supplier_map or {},
            "paymentTerms": inq.payment_terms,
            "confirmedAt": now_str(),
        },
        "quotations": [quotation_to_schema(q, db).model_dump() for q in quotations],
    }
    return QuotationSnapshot(
        id=gen_id("snap"),
        inquiry_id=inq.id,
        inquiry_code=inq.code,
        snapshot=frozen,
        created_at=now_str(),
        created_by=user.id if user else None,
        created_by_name=user.name if user else None,
    )


@router.get("/{inquiry_id}/snapshots", response_model=list[QuotationSnapshotSchema])
def list_quotation_snapshots(
    inquiry_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """返回询价单的报价快照列表（定标确认后冻结，不可变）。"""
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    rows = db.query(QuotationSnapshot).filter(
        QuotationSnapshot.inquiry_id == inquiry_id,
    ).order_by(QuotationSnapshot.created_at.desc()).all()
    return [
        QuotationSnapshotSchema(
            id=r.id,
            inquiryId=r.inquiry_id,
            inquiryCode=r.inquiry_code,
            createdAt=r.created_at,
            createdBy=r.created_by,
            createdByName=r.created_by_name,
            snapshot=r.snapshot or {},
        )
        for r in rows
    ]


def _export_dataset(db: Session, inq: Inquiry) -> dict:
    """构建导出数据（询价摘要 + 已提交报价）。"""
    quotations = db.query(Quotation).filter(Quotation.inquiry_id == inq.id).all()
    items = [{
        "name": it.name, "code": it.code, "spec": it.spec, "unit": it.unit,
        "quantity": it.quantity, "targetPrice": float(it.target_price) if it.target_price is not None else None,
    } for it in inq.items]
    quotes = []
    for q in quotations:
        if q.status != "SUBMITTED":
            continue
        quotes.append({
            "supplierId": q.supplier_id,
            "supplierName": q.supplier_name,
            "totalAmount": float(q.total_amount),
            "submittedAt": q.submitted_at,
            "items": [{
                "inquiryItemId": qi.inquiry_item_id,
                "unitPrice": float(qi.unit_price),
                "taxRate": float(qi.tax_rate),
                "taxIncludedTotal": float(qi.tax_included_total),
                "deliveryDays": qi.delivery_days,
                "warrantyMonths": qi.warranty_months,
                "paymentTerms": qi.payment_terms,
                "techDeviation": qi.tech_deviation,
                "commercialDeviation": qi.commercial_deviation,
            } for qi in q.items],
        })
    return {
        "inquiry": {
            "code": inq.code, "subject": inq.subject, "currency": inq.currency,
            "status": inq.status, "ownerName": inq.owner_name,
            "paymentTerms": inq.payment_terms, "deadline": inq.deadline,
        },
        "items": items,
        "quotations": quotes,
    }


def _export_xlsx(dataset: dict) -> bytes:
    """用 openpyxl 生成 Excel（字节流）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    # Sheet1 询价概览
    ws1 = wb.active
    ws1.title = "询价概览"
    inq = dataset["inquiry"]
    heads = ["询价编号", "主题", "币种", "状态", "负责人", "付款条件", "截止时间"]
    ws1.append(heads)
    for c in ws1[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E8E8E8")
    ws1.append([inq["code"], inq["subject"], inq["currency"], inq["status"], inq["ownerName"], inq["paymentTerms"], inq["deadline"]])

    # Sheet2 物料明细
    ws2 = wb.create_sheet("物料明细")
    ws2.append(["物料名称", "物料编码", "规格", "单位", "数量", "目标价"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for it in dataset["items"]:
        ws2.append([it["name"], it["code"], it["spec"], it["unit"], it["quantity"], it["targetPrice"]])

    # Sheet3 报价对比
    ws3 = wb.create_sheet("报价对比")
    ws3.append(["供应商", "报价总额", "报价时间", "币种"])
    for c in ws3[1]:
        c.font = Font(bold=True)
    for q in dataset["quotations"]:
        ws3.append([q["supplierName"], q["totalAmount"], q["submittedAt"], inq["currency"]])
    # 报价明细
    ws4 = wb.create_sheet("报价明细")
    ws4.append(["供应商", "物料编码", "单价", "税率", "含税总额", "交期(天)", "质保(月)", "技术偏离", "商务偏离"])
    for c in ws4[1]:
        c.font = Font(bold=True)
    for q in dataset["quotations"]:
        for qi in q["items"]:
            ws4.append([q["supplierName"], qi["inquiryItemId"], qi["unitPrice"], qi["taxRate"], qi["taxIncludedTotal"], qi["deliveryDays"], qi["warrantyMonths"], qi["techDeviation"] or "", qi["commercialDeviation"] or ""])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_pdf(dataset: dict) -> bytes:
    """用 reportlab 生成 PDF（字节流）。"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    import io
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    inq = dataset["inquiry"]
    story = []
    story.append(Paragraph(f"询价单 {inq['code']} - {inq['subject']}", styles['Title']))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"币种: {inq['currency']} | 状态: {inq['status']} | 负责人: {inq['ownerName']} | 付款条件: {inq['paymentTerms']} | 截止: {inq['deadline']}", styles['BodyText']))
    story.append(Spacer(1, 4 * mm))

    # 报价对比表
    header = ["供应商", "报价总额", "报价时间"]
    rows = [header]
    for q in dataset["quotations"]:
        rows.append([q["supplierName"], f"{q['totalAmount']:.2f}", q["submittedAt"] or ""])
    if len(rows) == 1:
        rows.append(["-", "-", "-"])
    t = Table(rows, colWidths=[60 * mm, 40 * mm, 60 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


@router.post("/{inquiry_id}/export")
def export_inquiry(
    inquiry_id: str,
    body: ExportRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """服务端生成询价详情/比价结果的 PDF 或 Excel，返回文件流。

    不依赖浏览器当前页面状态；scope 为 compare（默认）或 inquiry，两者均基于报价数据。
    """
    inq = db.query(Inquiry).filter(Inquiry.id == inquiry_id).first()
    if inq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="询价单不存在")
    require_inquiry_access(user, inq)
    fmt = (body.format or "xlsx").lower()
    if fmt not in ("pdf", "xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="format 仅支持 pdf / xlsx")
    dataset = _export_dataset(db, inq)
    if fmt == "pdf":
        content = _export_pdf(dataset)
        media = "application/pdf"
        ext = "pdf"
    else:
        content = _export_xlsx(dataset)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    filename = f"inquiry-{inq.code}.{ext}"
    return StreamingResponse(
        iter([content]),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
